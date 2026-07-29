# -*- coding: utf-8 -*-
"""
app.py — Streamlit

Upload em lote de vários trios de PDF (*_PDF.pdf, *_RPS.pdf, *_NF.pdf),
agrupamento automático, extração e geração de uma planilha Excel
consolidada (uma linha por alíquota, somando RPS + NF de cada trio).
"""
import io
import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from embratel_extractor import process_files, OUTPUT_COLUMNS, group_trios, classify_file

st.set_page_config(page_title="Extrator de Faturas EMBRATEL", layout="wide")


def build_excel_bytes(rows) -> bytes:
    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    body_font = Font(name="Arial")

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            if OUTPUT_COLUMNS[col_idx - 1] == "Valor":
                cell.number_format = "#,##0.00"
            if OUTPUT_COLUMNS[col_idx - 1] == "Tipo":
                cell.number_format = "0.00"

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        values = df[col_name].astype(str) if len(df) else pd.Series([], dtype=str)
        max_len = max([len(col_name)] + [len(v) for v in values])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


st.title("📄 Extrator de Faturas EMBRATEL/Claro — ArcelorMittal")
st.write(
    "Envie os PDFs de **um ou vários trios** de fatura (boleto `_PDF`, "
    "recibos `_RPS` e notas fiscais `_NF`). Os arquivos são agrupados "
    "automaticamente pelo nome — não precisa organizar em pastas."
)

uploaded_files = st.file_uploader(
    "Arquivos PDF (pode selecionar vários trios de uma vez)",
    type=["pdf"],
    accept_multiple_files=True,
)

if uploaded_files:
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_paths = []
        unrecognized = []
        for uf in uploaded_files:
            if classify_file(uf.name) is None:
                unrecognized.append(uf.name)
                continue
            p = Path(tmpdir) / uf.name
            p.write_bytes(uf.getbuffer())
            tmp_paths.append(p)

        if unrecognized:
            st.warning(
                "Arquivo(s) ignorado(s) — nome não termina em `_PDF.pdf`, "
                f"`_RPS.pdf` ou `_NF.pdf`: {', '.join(unrecognized)}"
            )

        trios = group_trios(tmp_paths)
        st.subheader(f"🔎 {len(trios)} trio(s) identificado(s)")

        incompletos = {k: v for k, v in trios.items() if len(v) < 3}
        if incompletos:
            st.warning(
                "Trio(s) incompleto(s) (faltando _PDF, _RPS ou _NF) — "
                "serão processados parcialmente:\n\n"
                + "\n".join(f"- `{k}` → tem: {', '.join(v.keys())}" for k, v in incompletos.items())
            )

        if st.button("▶️ Processar e gerar planilha", type="primary"):
            with st.spinner("Extraindo dados dos PDFs... (pode demorar um pouco em PDFs escaneados, que usam OCR)"):
                rows, trio_results = process_files(tmp_paths)

            st.subheader("✅ Resultado por trio")
            st.caption(
                "O total do boleto pode ser maior que o total agregado (RPS + NF) "
                "quando há juros/multa cobrados no boleto — isso é esperado e não "
                "é tratado como erro."
            )
            for tr in trio_results:
                icon = "✅" if not tr.avisos else "⚠️"
                with st.expander(f"{icon} {tr.key}  —  {len(tr.rows)} linha(s)"):
                    cols = st.columns(2)
                    if tr.total_boleto is not None:
                        cols[0].metric("Total do boleto", f"R$ {tr.total_boleto:,.2f}")
                    if tr.total_agregado is not None:
                        cols[1].metric("Total agregado (RPS + NF)", f"R$ {tr.total_agregado:,.2f}")
                    for aviso in tr.avisos:
                        st.warning(aviso)
                    if tr.rows:
                        st.dataframe(pd.DataFrame(tr.rows), use_container_width=True)

            if rows:
                df_final = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
                st.subheader(f"📊 Planilha consolidada — {len(df_final)} linha(s) no total")
                st.dataframe(df_final, use_container_width=True)

                excel_bytes = build_excel_bytes(rows)
                st.download_button(
                    "⬇️ Baixar planilha Excel consolidada",
                    data=excel_bytes,
                    file_name="EMBRATEL_faturas_consolidado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.error("Nenhuma linha foi gerada. Verifique os avisos acima.")
else:
    st.info("Aguardando upload dos arquivos PDF.")

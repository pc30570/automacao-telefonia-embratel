# -*- coding: utf-8 -*-
"""
gerar_planilha.py

Uso:
    python gerar_planilha.py <pasta_com_pdfs> <arquivo_saida.xlsx>

Varre a pasta em busca de trios de arquivos *_PDF.pdf / *_RPS.pdf / *_NF.pdf,
processa todos e gera uma única planilha Excel consolidada, no formato do
modelo EMBRATEL_COSS.
"""
import sys
import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from embratel_extractor import process_files, OUTPUT_COLUMNS

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")


def build_excel(rows, output_path: Path):
    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    body_font = Font(name="Arial")

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            if OUTPUT_COLUMNS[col_idx - 1] == "Valor":
                cell.number_format = "#,##0.00"
            if OUTPUT_COLUMNS[col_idx - 1] == "Tipo":
                cell.number_format = "0.00"

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        max_len = max([len(col_name)] + [len(str(v)) for v in df[col_name].astype(str)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def main():
    if len(sys.argv) != 3:
        print("Uso: python gerar_planilha.py <pasta_com_pdfs> <arquivo_saida.xlsx>")
        sys.exit(1)

    input_dir = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    pdf_paths = sorted(input_dir.glob("*.pdf"))
    if not pdf_paths:
        print(f"Nenhum PDF encontrado em {input_dir}")
        sys.exit(1)

    rows, trio_results = process_files(pdf_paths)

    print(f"\n{len(trio_results)} trio(s) encontrado(s):")
    for tr in trio_results:
        status = "OK" if not tr.avisos else "COM AVISOS"
        print(f"  - {tr.key}  [{status}]  {len(tr.rows)} linha(s)")
        for aviso in tr.avisos:
            print(f"      ! {aviso}")

    if not rows:
        print("\nNenhuma linha foi gerada. Verifique os avisos acima.")
        sys.exit(1)

    build_excel(rows, output_path)
    print(f"\nPlanilha gerada em: {output_path}  ({len(rows)} linha(s) no total)")


if __name__ == "__main__":
    main()
